"""
Herramientas para acceso a datos.
"""

import os, json, asyncio, logging
from datetime import datetime
from typing import Optional, List, Dict, Any
import pyodbc
import mysql.connector
import chainlit as cl

from QAgent.tools.base_tool import BaseTool
from QAgent.config.config_manager import config
from QAgent.utils.logging_utils import notify_error, get_random_response

from sqlalchemy import create_engine, text
from sqlalchemy.exc import DBAPIError, SQLAlchemyError
from sqlalchemy.pool import QueuePool
from sqlalchemy.ext.asyncio import create_async_engine, AsyncEngine
from sqlalchemy.engine import CursorResult
from sqlalchemy.pool.impl import AsyncAdaptedQueuePool

import pandas as pd
import json
from azure.identity import AzureCliCredential
import urllib
import struct
import re
from urllib.parse import quote_plus
from tenacity import retry, stop_after_attempt, wait_exponential


logger = logging.getLogger(__name__)

#=======================================================================================#

def _collect_error_chain(err: Exception) -> str:
    """Concatena mensajes desde err, su .orig (SQLAlchemy), __cause__ y __context__."""
    seen = set()
    parts = []

    def add_exc(e):
        if e and id(e) not in seen:
            seen.add(id(e))
            # Mensaje primario
            parts.append(str(e).strip())
            # Si es un DBAPIError, intenta extraer .orig (pyodbc/cx_*)
            if isinstance(e, DBAPIError) and getattr(e, "orig", None):
                parts.append(str(e.orig).strip())
            # pyodbc suele traer tuplas en .args
            if getattr(e, "args", None):
                try:
                    parts.extend([str(a).strip() for a in e.args if a])
                except Exception:
                    pass

    # Camina la cadena
    cur = err
    while cur:
        add_exc(cur)
        # SQLAlchemy DBAPIError -> orig
        if isinstance(cur, DBAPIError) and getattr(cur, "orig", None):
            add_exc(cur.orig)
        # Causas/Contexto encadenado
        if getattr(cur, "__cause__", None) and cur.__cause__ is not cur:
            cur = cur.__cause__
            continue
        if getattr(cur, "__context__", None) and cur.__context__ is not cur:
            cur = cur.__context__
            continue
        break

    # Deja algo compacto (evita duplicados exactos)
    compact = []
    for p in parts:
        if p and p not in compact:
            compact.append(p)
    return " | ".join(compact)

def _classify_sql_error(full_msg: str) -> str:
    """Devuelve un mensaje comprensible para la IA/usuario según patrones conocidos."""
    lower = full_msg.lower()

    # Patrones muy comunes de SQL Server
    if "invalid object name" in lower or "object not found" in lower:
        return ("Error: la tabla o vista indicada no existe en SQL Server. "
                "Verifica el nombre (con esquema) y vuelve a intentarlo.")
    if "invalid column name" in lower:
        return ("Error: una de las columnas referenciadas no existe. "
                "Revisa los nombres de columnas y vuelve a intentarlo.")
    if "login failed" in lower or "not authorized" in lower or "permission" in lower or "401" in lower:
        return ("Error de permisos/autenticación al acceder a SQL Server. "
                "Solicita al administrador el acceso requerido.")
    if "deadlock victim" in lower or "1205" in lower:
        return ("Error: deadlock detectado (1205). Reintenta la operación y/o ajusta la transacción "
                "(índices, orden de actualización, granularidad).")
    if "lock request time out" in lower or "1222" in lower or "timeout" in lower or "timed out" in lower:
        return ("Timeout/bloqueo al consultar la base de datos. Optimiza la consulta, revisa índices "
                "o vuelve a intentarlo.")
    if "string or binary data would be truncated" in lower or "2628" in lower:
        return ("Error: datos demasiado largos para el tipo de columna (truncamiento). "
                "Ajusta longitudes o valida los datos de entrada.")
    if "cannot insert the value null" in lower or "515" in lower:
        return ("Error: intento de insertar NULL en una columna que no lo permite. "
                "Valida los campos obligatorios.")
    if "violation of unique key constraint" in lower or "cannot insert duplicate key" in lower or "2627" in lower or "2601" in lower:
        return ("Error: clave duplicada (índice único). Ajusta el valor o la lógica de inserción.")
    if "conversion failed" in lower:
        return ("Error de conversión de tipos (por ejemplo, string→numérico/fecha). "
                "Valida y castea correctamente los datos.")
    if "divide by zero" in lower:
        return ("Error: división por cero en la consulta. Ajusta el cálculo o usa NULLIF/CASE.")

    # El tuyo: “No corresponding transaction found. (111214)”
    if "no corresponding transaction found" in lower or "111214" in lower:
        return ("Error de transacción: se intentó completar/confirmar una transacción inexistente. "
                "Revisa el manejo de begin/commit/rollback y el uso de conexiones/pooled sessions.")

    # Fallback genérico
    return ("Ocurrió un error de base de datos al ejecutar la consulta. "
            "Revisa la sintaxis, permisos y conectividad.")

async def _handle_db_exception(err: Exception) -> str:
    """
    Construye un mensaje 'amigable' + adjunta detalle técnico (compacto).
    También registra TODO en logs.
    """
    full = _collect_error_chain(err)
    user_friendly = _classify_sql_error(full)
    # Log completo para troubleshooting
    logger.error("DB ERROR (full chain): %s", full)
    # Devuelve mensaje para la IA/usuario + una línea de detalle técnico (recortada)
    tail = full[:600]  # limita tamaño
    return f"{user_friendly}\n\nDetalle técnico: {tail}"

#=======================================================================================#
#=======================================================================================#
#=======================================================================================#

_lit_re = re.compile(r"'((?:[^']|'')*)'")  # captura literal SQL con '' escapadas

def _parameterize_percent_literals(sql: str):
    """
    Reemplaza literales con % por :litN y devuelve (sql_nuevo, params).
    Mantiene otros literales intactos.
    """
    params = {}
    out = []
    last = 0
    idx = 0
    for m in _lit_re.finditer(sql):
        start, end = m.span()
        inner = m.group(1)
        out.append(sql[last:start])
        if "%" in inner:
            # des-escapar '' -> '
            val = inner.replace("''", "'")
            key = f"lit_{idx}"
            params[key] = val
            out.append(f":{key}")
            idx += 1
        else:
            # dejar literal tal cual
            out.append(sql[start:end])
        last = end
    out.append(sql[last:])
    return "".join(out), params

TABLAS_VALIDAS = {
    "base_envios",
    "cep",
    "clientes"
}

class GetDataMySQLAsync:
    """
    Acceso async a MySQL con:
      - AsyncEngine (mysql+aiomysql | mysql+asyncmy)
      - Pool async
      - Pandas vía conn.run_sync(...)
      - Exportación a Excel en hilo aparte si hay > 20 filas
      - Manejo de errores robusto (mensajes específicos de MySQL)
    """

    _engine: Optional[AsyncEngine] = None

    def __init__(
        self,
        *,
        public_root: str = "public/storage/excel",
        base_download_url: str = config.get("URL_ARCHIVO"),
        pool_size: int = 10,
        max_overflow: int = 20,
        pool_timeout: int = 30,
        pool_recycle: int = 3600,
        echo: bool = False,
        driver: str = "aiomysql",  # "aiomysql" o "asyncmy"
        ssl_ca: Optional[str] = None,  # ruta a CA si corresponde
        ssl_cert: Optional[str] = None,
        ssl_key: Optional[str] = None,
    ):
        self.public_root = public_root
        self.base_download_url = base_download_url.rstrip("/") + "/"
        self.pool_size = pool_size
        self.max_overflow = max_overflow
        self.pool_timeout = pool_timeout
        self.pool_recycle = pool_recycle
        self.echo = echo
        self.driver = driver
        self.ssl_ca = ssl_ca
        self.ssl_cert = ssl_cert
        self.ssl_key = ssl_key

        os.makedirs(self.public_root, exist_ok=True)


      
    # -----------------------------
    # Engine / Pool (singleton)
    # -----------------------------
    def get_engine(self) -> AsyncEngine:
        if self.__class__._engine is None:
            HOST     = config.get("DB_HOST", "127.0.0.1")
            PORT     = int(config.get("DB_MYSQL_PORT", 3306))
            DATABASE = config.get("DB_NAME")
            USERNAME = config.get("DB_USER")
            PASSWORD = config.get("DB_PASSWORD")
            
            u = quote_plus(USERNAME or "")
            p = quote_plus(PASSWORD or "")
            
            logger.info(f"DATABASE:{DATABASE}, USERNAME:{USERNAME}")
           

            # Driver: "aiomysql" (compatible) o "asyncmy" (más performante)
            # Nota: charset=utf8mb4 recomendado; autocommit=true útil para SELECT simples
            dialect = "aiomysql" if self.driver == "aiomysql" else "asyncmy"
            if PASSWORD:
                conn_url = f"mysql+aiomysql://{u}:{p}@{HOST}:{PORT}/{DATABASE}"
            else:
                conn_url = f"mysql+aiomysql://{u}@{HOST}:{PORT}/{DATABASE}"
            
            

            connect_args = {
                # timeouts razonables; algunos drivers usan otros nombres
                "connect_timeout": 30,
            }

            # SSL opcional (si MySQL lo requiere)
            if self.ssl_ca or self.ssl_cert or self.ssl_key:
                # Para aiomysql: pasar dict ssl={"ca":..., "cert":..., "key":...}
                # Para asyncmy: ssl también es soportado con dict compatible.
                ssl_dict = {}
                if self.ssl_ca:   ssl_dict["ca"] = self.ssl_ca
                if self.ssl_cert: ssl_dict["cert"] = self.ssl_cert
                if self.ssl_key:  ssl_dict["key"] = self.ssl_key
                connect_args["ssl"] = ssl_dict

            self.__class__._engine = create_async_engine(
                conn_url,
                echo=self.echo,
                poolclass=AsyncAdaptedQueuePool,
                pool_size=self.pool_size,
                max_overflow=self.max_overflow,
                pool_timeout=self.pool_timeout,
                pool_recycle=self.pool_recycle,  # evita conexiones zombies
                pool_pre_ping=True,              # detecta conexiones caídas
                connect_args=connect_args,
            )
            logger.info(f"AsyncEngine MySQL ({dialect}) inicializado.")
        return self.__class__._engine

    # ---------------------------------
    # Exportador Excel (no bloqueante)
    # ---------------------------------
    async def _generar_excel_async(self, rows: List[Dict[str, Any]]) -> str:
        fecha = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        filename = f"reporte_{fecha}.xlsx"
        dest_path = os.path.join(self.public_root, filename)
        
        print(f"--> 1 Dest_path : {dest_path}")

        def _write_excel_sync():
            import openpyxl
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active

            if not rows:
                wb.save(dest_path)
                return

            columnas = list(rows[0].keys())
            # encabezados
            for c, col in enumerate(columnas, 1):
                ws.cell(row=1, column=c, value=col)

            # datos
            for r, fila in enumerate(rows, 2):
                for c, col in enumerate(columnas, 1):
                    val = fila.get(col)
                    # strings para objetos tipo fecha/decimal
                    if hasattr(val, "isoformat"):
                        val = val.isoformat()
                    ws.cell(row=r, column=c, value=val)

            # ancho columnas
            for c, col in enumerate(columnas, 1):
                longs = [len(str(col))]
                for item in rows:
                    v = item.get(col)
                    if v is not None:
                        longs.append(len(str(v)))
                ws.column_dimensions[get_column_letter(c)].width = min(max(longs) + 2, 50)

            wb.save(dest_path)

        await asyncio.to_thread(_write_excel_sync)

        rel_path = os.path.relpath(dest_path, start=".").replace("\\", "/")
        # si no configuras URL base, devuelvo ruta relativa como fallback
        url = f"{self.base_download_url}{rel_path}" if self.base_download_url else rel_path
        logger.info(f"Excel generado: {url}")
        return url

    # -----------------------------
    # Ejecución de consulta
    # -----------------------------
    @retry(stop=stop_after_attempt(2), wait=wait_exponential(multiplier=1, max=4), reraise=True)
    async def execute(self, sql: str) -> str:
        """
        Devuelve JSON con:
          - Primer item: {"Instrucción adicional": "..."}
          - Luego hasta 20 filas (preview)
          - Si >20 filas, genera Excel y agrega link en la instrucción
        """
        engine = self.get_engine()
        logger.info(f"Ejecutando SQL MySQL ({len(sql)} chars)")

        try:
            async with engine.begin() as conn:
                # pandas.read_sql es sync -> run_sync evita bloquear el event loop
                
                sql_safe, params = _parameterize_percent_literals(sql)
                
                df: pd.DataFrame = await conn.run_sync(
                     lambda sc: pd.read_sql(text(sql_safe), con=sc, params=params)
                )

            total_rows = len(df.index)
            logger.info(f"Filas recuperadas: {total_rows}")

            preview_count = 20
            preview_df = df.head(preview_count)
            preview = preview_df.to_dict(orient="records")

            payload: List[Dict[str, Any]] = []

            if total_rows == 0:
                payload.append({
                    "Instrucción adicional": (
                        "La consulta no devolvió resultados. "
                        "Indica explícitamente que no hay filas y no inventes datos."
                    )
                })
            elif total_rows > preview_count:
                records_full = df.to_dict(orient="records")
                excel_url = await self._generar_excel_async(records_full)
                print(f"--> 1 excel_url : {excel_url}")
                payload.append({
                    "Instrucción adicional": (
                        f"# Atención, esto es muy importante para el usuario:\n"
                        f"- El resultado tiene {total_rows} filas. "
                        f"**Despliega solo las primeras {preview_count} filas** y "
                        f"**entrega el archivo completo en este link**: "
                        f"[Descargar Excel]({excel_url}) (renderiza el link en formato MD)."
                    )
                })
            else:
                payload.append({
                    "Instrucción adicional": (
                        f"El resultado tiene {total_rows} filas. "
                        f"Entrega una tabla en **Markdown** con todas las filas."
                    )
                })

            payload.extend(preview)

            json_resultado = json.dumps(payload, ensure_ascii=False, default=str)
            tamano_bytes = len(json_resultado.encode("utf-8"))
            logger.info(f"📦 Tamaño de respuesta: {tamano_bytes} bytes")

            if tamano_bytes > 500_000:
                logger.warning("⚠️ Respuesta supera los 500 KB")
                return (
                    "Respuesta demasiado grande. "
                    "Debes paginar, resumir o limitar con LIMIT/OFFSET. Inténtalo nuevamente."
                )

            return json_resultado

        except DBAPIError as err:
            # Mensaje amigable al usuario y log
            try:
                response = get_random_response("error")
                await cl.Message(response).send()
            except Exception:
                pass

            msg = str(err).strip()
            logger.error(f"DBAPIError MySQL: {msg}")
            lower = msg.lower()

            # Errores típicos de MySQL
            if "unknown table" in lower or "table doesn't exist" in lower or "doesn't exist" in lower:
                return ("Error: la tabla o vista indicada no existe en MySQL. "
                        "Verifica el nombre (y esquema si aplica) y vuelve a intentarlo.")

            if "unknown column" in lower or "column not found" in lower or "invalid column" in lower:
                return ("Error: una de las columnas referenciadas no existe. "
                        "Revisa los nombres de columnas y vuelve a intentarlo.")

            if "access denied" in lower or "not authorized" in lower or "permission" in lower:
                return ("Error de permisos/autenticación al acceder a MySQL. "
                        "Solicita al administrador el acceso requerido.")

            if "lock wait timeout" in lower:
                return ("Timeout por espera de bloqueo en MySQL (LOCK WAIT TIMEOUT). "
                        "Intenta reducir el alcance de la consulta o reintentar.")

            if "deadlock" in lower:
                return ("Se detectó un DEADLOCK en MySQL. "
                        "Vuelve a ejecutar o ajusta la transacción/consulta para minimizar bloqueos.")

            if "packet too large" in lower:
                return ("El resultado excede el tamaño de paquete permitido en MySQL. "
                        "Reduce columnas/filas o ajusta 'max_allowed_packet'.")

            if "timeout" in lower or "timed out" in lower or "lost connection" in lower:
                return ("Timeout o pérdida de conexión con MySQL. "
                        "Optimiza la consulta o vuelve a intentarlo.")

            # Validación ligera con TABLAS_VALIDAS si las usas
            try:
                sql_upper = sql.upper()
                alguna_valida = any(tabla in sql_upper for tabla in TABLAS_VALIDAS)
                alguna_invalida = False
                # ejemplo simple: detectar backticks o esquema `ia.`
                for palabra in sql_upper.replace("`", "").split():
                    if palabra.startswith('IA.') and palabra not in TABLAS_VALIDAS:
                        alguna_invalida = True
                        break
                if alguna_invalida or not alguna_valida:
                    return ("Error: la tabla o vista indicada no existe en MySQL. "
                            f"Tablas válidas: {', '.join(sorted(TABLAS_VALIDAS))}" if TABLAS_VALIDAS else
                            "Error: la tabla o vista indicada no existe en MySQL.")
            except Exception:
                pass

            return f"Ocurrió un error de base de datos al ejecutar la consulta. Detalle: {msg}"

        except SQLAlchemyError as err:
            try:
                response = get_random_response("error")
                await cl.Message(response).send()
            except Exception:
                pass

            msg = str(err).strip()
            logger.error(f"SQLAlchemyError MySQL: {msg}")
            return ("Ocurrió un error interno al preparar/ejecutar la consulta. "
                    f"Detalle: {msg}")

        except Exception as err:
            try:
                response = get_random_response("error")
                await cl.Message(response).send()
            except Exception:
                pass

            logger.exception("Error inesperado ejecutando la consulta (MySQL)")
            return f"Ocurrió un error inesperado. Detalle: {err}"

#============================================================================================#
#============================================================================================#
#============================================================================================#

class GetDataLocalMySQLTool(BaseTool):
    """
    Herramienta para ejecutar consultas SQL y obtener datos.
    """
    
    async def execute(self, consulta: str) -> str:
        """
        Ejecuta una consulta SQL y devuelve los resultados como JSON
        
        Args:
            consulta: Consulta SQL a ejecutar
            
        Returns:
            Resultados de la consulta en formato JSON
        """
        db_connection = None
        cursor = None
        try:
            user = cl.user_session.get("user")
            logger.info("---> USUARIO DDBB---\n")
            logger.info(F"{user.identifier}")
            logger.info("---> FIN USUARIO DDBB---\n")
            
            if user.identifier =="quinta@qagent.cl":
                logger.info("---> Base de datos quinta ---\n")
                db_connection = mysql.connector.connect(
                    host='localhost',
                    user='root',
                    password='',
                    database='quinta',
                )
            
            
            elif user.identifier =="correos@qagent.cl": 
                logger.info("---> Base de datos correos de chile ---\n")   
                db_connection = mysql.connector.connect(
                    host='localhost',
                    user='root',
                    password='',
                    database='correos_chile',
                )
            
            
            else:
                 db_connection = mysql.connector.connect(
                    host='localhost',
                    user='root',
                    password='',
                    database='cpp',
                ) 
           
            
           
            cursor = db_connection.cursor(dictionary=True) 
            logger.info(f"Ejecutando consulta: {consulta}")
            
            # Ejecutar la consulta
            cursor.execute(consulta)
            resultados = cursor.fetchall()            
            filas = len(resultados)
            
            if filas > 100:                
                url_archo_excel = await generar_excel_desde_mysql(resultados)    
                resultados.insert(0, {"Instrucción adicional": f" # Atención, esto es muy importante para el usuario: - El resultado tiene {filas} filas. **Despliega solo las primeras 15 filas e informa al usuario que puede descargar el archivo excel con todos los datos en el siguiente link [Descargar Excel]({url_archo_excel}) renderiza el link en formato md.**"})        
            elif filas > 50:
                url_archo_excel = await generar_excel_desde_mysql(resultados)    
                resultados.insert(0, {"Instrucción adicional": f" # Atención, esto es muy importante para el usuario: - El resultado tiene {filas} filas. **Despliega solo las primeras 15 filas e informa al usuario que puede descargar el archivo excel con todos los datos en el siguiente link [Descargar Excel]({url_archo_excel}) renderiza el link en formato md.**"})        
            elif filas > 20:
                resultados.insert(0, {"Instrucción adicional": f"El resultado tiene {filas} filas. Debes utilizar la herramienta createDataFrame() para paginar los resultados, no formato markdown. **debe estar en formato dict serializado en JSON. Nunca lo envíes como un string anidado ni como tabla Markdown y limpia los campos null.** "})
            else:
                resultados.insert(0, {"Instrucción adicional": f"El resultado tiene {filas} filas. Recuerda que para entregar listas de menos de 20 registros debes utilizar formato markdown"})
            
            
            json_resultados = json.dumps(resultados, default=str)
            
            logger.info(f"Q caracteres respuesta: {len(json_resultados)}")
            return json_resultados
            
        except mysql.connector.Error as err:
            error_message = str(err)
            
            # Gestión de errores
            modulo = "data_tools"
            funcion = "getdata"
            #await notify_error(error_message, modulo, funcion)
            
            response = get_random_response("error")
            await cl.Message(response).send()
            
            if "Unknown column" in error_message:
                return f"Cometiste el siguiente error en tu consulta:{err}. Por favor revisa el nombre de las columnas en tu base de conocimiento y otorga la respuesta correcta"
            else:
                return f"Cometiste el siguiente error en tu consulta:{err}. Por favor replantea la consulta y otorga la respuesta correcta"
        
        finally:
            # Asegura que el cursor y la conexión se cierran
            if cursor:
                cursor.close()
            if db_connection and db_connection.is_connected():
                db_connection.close()

class GetDataAzureSQLServerAD(BaseTool):
    
   
    """
    Herramienta para ejecutar consultas SQL SERVER en Azure usando AAD.
    Usa token AAD de Azure CLI y engine con pool global.
    """
    _engine = None

    @classmethod
    def get_engine(cls):
        if cls._engine is None:
            SERVER   = config.get("DB_ASQLS_SERVER")
            DATABASE = config.get("DB_ASQLS_DATABASE")
            DRIVER   = config.get("DB_ASQLS_DRIVER")

            logger.info("🔐 Obteniendo token desde Azure CLI...")
            os.environ["PATH"] += os.pathsep + "/usr/bin"
            credential = AzureCliCredential()
            token = credential.get_token("https://database.windows.net/.default").token.encode("utf-16-le")
            token_struct = struct.pack(f"<I{len(token)}s", len(token), token)
            SQL_COPT_SS_ACCESS_TOKEN = 1256

            conn_str = (
                f"Driver={{{DRIVER}}};"
                f"Server={SERVER};"
                f"Database={DATABASE};"
                f"Encrypt=yes;"
                f"TrustServerCertificate=no;"
            )
            quoted_conn_str = urllib.parse.quote_plus(conn_str)
            url = f"mssql+pyodbc:///?odbc_connect={quoted_conn_str}"

            cls._engine = create_engine(
                url,
                poolclass=QueuePool,
                pool_size=10,
                max_overflow=20,
                pool_timeout=30,
                pool_recycle=3600,
                pool_pre_ping=True,
                fast_executemany=True,
                echo=False,
                connect_args={
                    "attrs_before": {SQL_COPT_SS_ACCESS_TOKEN: token_struct},
                    "application_name": "Qagent_App",
                    "autocommit": False,
                },
            )
        return cls._engine

    async def execute(self, consulta: str) -> str:
        try:
            engine = self.get_engine()
            logger.info(f"Ejecutando consulta: {consulta}")

            with engine.connect() as conn:
                try:
                    logger.info("✅ Conexión abierta, ejecutando consulta con pd.read_sql desde raw pyodbc connection...")
                    
                    # Usar conexión pyodbc directamente
                    raw_conn = conn.connection.connection  # 👈 no el de SQLAlchemy
                    
                    df_resultado = pd.read_sql(consulta, raw_conn)
                    logger.info("✅ Consulta ejecutada correctamente.")
                except Exception as e:
                    logger.error(f"❌ Error durante pd.read_sql: {e}")
                    raise

            resultados = df_resultado.to_dict(orient="records")
            filas = len(resultados)


            if filas > 100:
                url_archo_excel = await generar_excel_desde_mysql(resultados)
                resultados.insert(0, {
                    "Instrucción adicional": (
                        f"# Atención, esto es muy importante para el usuario: "
                        f"- El resultado tiene {filas} filas. "
                        f"**Despliega solo las primeras 20 filas e informa al usuario que puede descargar el archivo excel con todos los datos en el siguiente link [Descargar Excel]({url_archo_excel}) renderiza el link en formato md.**"
                    )
                })
            elif filas >= 50:
                url_archo_excel = await generar_excel_desde_mysql(resultados)
                resultados.insert(0, {
                    "Instrucción adicional": (
                        f"# Atención, esto es muy importante para el usuario: "
                        f"- El resultado tiene {filas} filas. "
                        f"**Despliega solo las primeras 20 filas e informa al usuario que puede descargar el archivo excel con todos los datos en el siguiente link [Descargar Excel]({url_archo_excel}) renderiza el link en formato md.**"
                    )
                })
            elif filas >= 15:
                url_archo_excel = await generar_excel_desde_mysql(resultados)
                resultados.insert(0, {
                    "Instrucción adicional": (
                        f"# Atención, esto es muy importante para el usuario: "
                        f"- El resultado tiene {filas} filas. "
                        f"**Despliega solo las primeras 20 filas e informa al usuario que puede descargar el archivo excel con todos los datos en el siguiente link [Descargar Excel]({url_archo_excel}) renderiza el link en formato md.**"
                    )
                })
            else:
                resultados.insert(0, {
                    "Instrucción adicional": (
                        f"El resultado tiene {filas} filas. Recuerda que para entregar listas de menos de 20 registros debes utilizar formato markdown"
                    )
                })

            json_resultado = json.dumps(resultados, default=str)
            tamano_bytes = len(json_resultado.encode("utf-8"))
            logger.info(f"Q Tamaño de respuesta: {tamano_bytes}")

            if tamano_bytes > 500_000:
                logger.warning("Respuesta supera los 500KB")
                return "Respuesta demasiado grande. Debes paginar, resumir o limitar con LIMIT. Por favor intenta nuevamente"
            else:
                return json_resultado

        except SQLAlchemyError as err:
            error_message = str(err)
            logger.error(f"Error {error_message}")
            response = get_random_response("error")
            await cl.Message(response).send()

            if "Invalid column name" in error_message:
                return f"Cometiste el siguiente error en tu consulta: {err}. Por favor revisa el nombre de las columnas en tu base de conocimiento y otorga la respuesta correcta"
            
            elif "timeout" in error_message.lower() or "HYT00" in error_message:
                return "La consulta se demoró demasiado y fue cancelada automáticamente. Por favor intenta optimizarla o dividirla en partes más pequeñas."

            else:
                return f"Cometiste el siguiente error en tu consulta: {err}. Por favor replantea la consulta y otorga la respuesta correcta"

class GetDataAzureSQLServer(BaseTool):
    """
    Herramienta para ejecutar consultas SQL SERVER y obtener datos.
    Usa un engine global para aprovechar el pool de conexiones.
    """
    # --- Engine global para la clase ---
    _engine = None

    @classmethod
    def get_engine(cls):
        if cls._engine is None:
            SERVER   = config.get("DB_ASQLS_SERVER")
            DATABASE = config.get("DB_ASQLS_DATABASE")
            USERNAME = config.get("DB_ASQLS_USERNAME")
            PASSWORD = config.get("DB_ASQLS_PASSWORD")
            DRIVER   = config.get("DB_ASQLS_DRIVER")

            url = (
                f"mssql+pyodbc://{USERNAME}:{PASSWORD}"
                f"@{SERVER}:1433/{DATABASE}"
                f"?driver={DRIVER.replace(' ', '+')}&TrustServerCertificate=yes&encrypt=yes"
            )
            
            
            cls._engine = create_engine(
                url,
                poolclass=QueuePool,
                pool_size=10,
                max_overflow=20,
                pool_timeout=30,
                pool_recycle=3600,
                pool_pre_ping=True,
                fast_executemany=True,
                echo=False,
                connect_args={
                    'application_name': 'Qagent_App',
                    'timeout': 30
                }
            )
        return cls._engine

    async def execute(self, consulta: str) -> str:
        try:
            engine = self.get_engine()
            logger.info(f"Ejecutando consulta: {consulta}")

            with engine.connect() as conn:
                df_resultado = pd.read_sql(consulta, conn)
            resultados = df_resultado.to_dict(orient="records")
            filas = len(resultados)           
            
            
            if filas > 20:                
                url_archo_excel = await generar_excel_desde_mysql(resultados)    
                resultados.insert(0, {"Instrucción adicional": f" # Atención, esto es muy importante para el usuario: - El resultado tiene {filas} filas. **Despliega solo las primeras 20 filas e informa al usuario que puede descargar el archivo excel con todos los datos en el siguiente link [Descargar Excel]({url_archo_excel}) renderiza el link en formato md.**"})        
            else:
                resultados.insert(0, {"Instrucción adicional": f"El resultado tiene {filas} filas. Recuerda que para entregar listas de menos de 20 registros debes utilizar formato markdown"})
            

            json_resultado = json.dumps(resultados, default=str)
            tamano_bytes = len(json_resultado.encode("utf-8"))
            logger.info(f"Q Tamaño de respuesta: {tamano_bytes}")

            if tamano_bytes > 500000:
                logger.warning("Respuesta supera los 500KB")
                return "Respuesta demasiado grande. Debes paginar, resumir o limitar con LIMIT. Por favor intenta nuevamente"
            else:
                return json_resultado

        except SQLAlchemyError as err:
            error_message = str(err)
            response = get_random_response("error")
            await cl.Message(response).send()

            if "Invalid column name" in error_message:
                return f"Cometiste el siguiente error en tu consulta: {err}. Por favor revisa el nombre de las columnas en tu base de conocimiento y otorga la respuesta correcta"
            else:
                return f"Cometiste el siguiente error en tu consulta: {err}. Por favor replantea la consulta y otorga la respuesta correcta"

class GetDataGCPSQLServer(BaseTool):
    """
    Herramienta para ejecutar consultas SQL SERVER en GCP y obtener datos.
    Usa un engine global para aprovechar el pool de conexiones.
    """
    _engine = None

    @classmethod
    def get_engine(cls):
        if cls._engine is None:
            SERVER   = config.get("DB_GCP_SQLS_SERVER")
            DATABASE = config.get("DB_CP_SQLS_DATABASE")
            USERNAME = config.get("DB_DB_GCP_SQLS_USERNAME")
            PASSWORD = config.get("DB_GCP_SQLS_PASSWORD")
            DRIVER   = config.get("DB_GCP_SQLS_DRIVER") 

            url = (
                f"mssql+pyodbc://{USERNAME}:{PASSWORD}"
                f"@{SERVER}:1433/{DATABASE}"
                f"?driver={DRIVER.replace(' ', '+')}"
                f"&TrustServerCertificate=yes"
                f"&encrypt=yes"
            )
            cls._engine = create_engine(
                url,
                poolclass=QueuePool,
                pool_size=10,
                max_overflow=20,
                pool_timeout=30,
                pool_recycle=3600,
                pool_pre_ping=True,
                fast_executemany=True,
                echo=False,
                connect_args={
                    'application_name': 'Qagent_GCP_App',
                    'timeout': 30
                }
            )
        return cls._engine

    async def execute(self, consulta: str) -> str:
        try:
            engine = self.get_engine()
            logger.info(f"Ejecutando consulta: {consulta}")

            with engine.connect() as conn:
                df_resultado = pd.read_sql(consulta, conn)
            resultados = df_resultado.to_dict(orient="records")
            filas = len(resultados)

            json_resultado = json.dumps(resultados, default=str)
            tamano_bytes = len(json_resultado.encode("utf-8"))
            logger.info(f"Q Tamaño de respuesta: {tamano_bytes}")

            if tamano_bytes > 500000:
                logger.warning("Respuesta supera los 500KB")
                return "Respuesta demasiado grande. Debes paginar, resumir o limitar con LIMIT. Por favor intenta nuevamente"
            else:
                return json_resultado

        except SQLAlchemyError as err:
            error_message = str(err)
            response = get_random_response("error")
            await cl.Message(response).send()

            if "Invalid column name" in error_message:
                return f"Cometiste el siguiente error en tu consulta: {err}. Por favor revisa el nombre de las columnas en tu base de conocimiento y otorga la respuesta correcta"
            else:
                return f"Cometiste el siguiente error en tu consulta: {err}. Por favor replantea la consulta y otorga la respuesta correcta"

class CreateDataFrameTool(BaseTool):
  
    async def execute(self, message: str, dataframe: str) -> str:
        try:
            import pandas as pd
            import json
            
            logger.info("==== Asi llego el dataframe ====")
            logger.info(dataframe)
            logger.info("===============================")

            # Parseo robusto
            if isinstance(dataframe, str):
                logger.info("Se intentó parsear 1 vez")
                parsed = json.loads(dataframe)
                if isinstance(parsed, str):
                    logger.info("Se intentó parsear 2 vez")
                    parsed = json.loads(parsed)
            else:
                logger.info("Se intentó parsear")
                parsed = dataframe
                
                
            logger.info("==== Asi termino el dataframe ====")
            logger.info(parsed)
            logger.info("===============================")     

            df = pd.DataFrame(parsed)
            elements = [cl.Dataframe(data=df, display="inline", name="Dataframe")]
            await cl.Message(content=message, elements=elements).send()
            return "Dataframe creado exitosamente y enviado al usuario."                     
           
            
        except Exception as err:
            logger.warning("No pudo generar el reporte de forma paginada")
            return f"""
            Error intentar crear al crear el  dataframe: {err}, recuerda el formato de la data debe ser un diccionario con el siguiente formato:
            Los valores deben estar en listas del mismo largo, por ejemplo:
            - "Nombre_columna_1": lista de strings
            - "Nombre_columna_2": lista de números
            Evita los null o reemplazalos por un valor por defecto, por ejemplo: "", "N/A", None, etc.
            Si continua generando error despliega en formato markdown
            """            

class GetDataSQLSLocal(BaseTool):
    """
    Herramienta para ejecutar consultas SQL SERVER en GCP y obtener datos.
    Usa un engine global para aprovechar el pool de conexiones.
    """
    _engine = None

    @classmethod
    def get_engine(cls):
        if cls._engine is None:
            SERVER   = "localhost"
            DATABASE = "qagent"
            DRIVER   = "ODBC Driver 18 for SQL Server"
                        
            url = (
                    f"mssql+pyodbc://@{SERVER}/{DATABASE}"
                    f"?driver={DRIVER.replace(' ', '+')}"
                    f"&Trusted_Connection=yes"
                    f"&TrustServerCertificate=yes"
                    f"&encrypt=no"
                )
            cls._engine = create_engine(
                 url,
                poolclass=QueuePool,
                pool_size=10,
                max_overflow=20,
                pool_timeout=30,
                pool_recycle=3600,
                pool_pre_ping=True,
                fast_executemany=True,
                echo=False,
                connect_args={
                    'application_name': 'Qagent_Local_App',
                    'timeout': 30
                }
            )
        return cls._engine

    async def execute(self, consulta: str) -> str:
        try:
            engine = self.get_engine()
            logger.info(f"Ejecutando consulta: {consulta}")

            with engine.connect() as conn:
                df_resultado = pd.read_sql(consulta, conn)
            resultados = df_resultado.to_dict(orient="records")
            filas = len(resultados)           
            
            
            if filas > 100:                
                url_archo_excel = await generar_excel_desde_mysql(resultados)    
                resultados.insert(0, {"Instrucción adicional": f" # Atención, esto es muy importante para el usuario: - El resultado tiene {filas} filas. **Despliega solo las primeras 20 filas e informa al usuario que puede descargar el archivo excel con todos los datos en el siguiente link [Descargar Excel]({url_archo_excel}) renderiza el link en formato md.**"})        
            elif filas >= 50:
                url_archo_excel = await generar_excel_desde_mysql(resultados)    
                resultados.insert(0, {"Instrucción adicional": f" # Atención, esto es muy importante para el usuario: - El resultado tiene {filas} filas. **Despliega solo las primeras 20 filas e informa al usuario que puede descargar el archivo excel con todos los datos en el siguiente link [Descargar Excel]({url_archo_excel}) renderiza el link en formato md.**"})        
            elif filas >= 15:               
                resultados.insert(0, {"Instrucción adicional": f"El resultado tiene {filas} filas. Debes utilizar la herramienta createDataFrame() para paginar los resultados, no formato markdown. **debe estar en formato dict serializado en JSON. Nunca lo envíes como un string anidado ni como tabla Markdown y limpia los campos null.** "})
            else:
                resultados.insert(0, {"Instrucción adicional": f"El resultado tiene {filas} filas. Recuerda que para entregar listas de menos de 20 registros debes utilizar formato markdown"})
            

            json_resultado = json.dumps(resultados, default=str)
            tamano_bytes = len(json_resultado.encode("utf-8"))
            logger.info(f"Q Tamaño de respuesta: {tamano_bytes}")

            if tamano_bytes > 500000:
                logger.warning("Respuesta supera los 500KB")
                return "Respuesta demasiado grande. Debes paginar, resumir o limitar con LIMIT. Por favor intenta nuevamente"
            else:
                return json_resultado

        except SQLAlchemyError as err:
            error_message = str(err)
            response = get_random_response("error")
            await cl.Message(response).send()

            if "Invalid column name" in error_message:
                return f"Cometiste el siguiente error en tu consulta: {err}. Por favor revisa el nombre de las columnas en tu base de conocimiento y otorga la respuesta correcta"
            else:
                return f"Cometiste el siguiente error en tu consulta: {err}. Por favor replantea la consulta y otorga la respuesta correcta"

class GetExplainSQL(BaseTool):
    """
    Herramienta para explicar  consultas SQL SERVER Local
    """
    _engine = None

    @classmethod
    def get_engine(cls):
        if cls._engine is None:
            SERVER   = "localhost"
            DATABASE = "qagent"
            DRIVER   = "ODBC Driver 18 for SQL Server"
                        
            url = (
                    f"mssql+pyodbc://@{SERVER}/{DATABASE}"
                    f"?driver={DRIVER.replace(' ', '+')}"
                    f"&Trusted_Connection=yes"
                    f"&TrustServerCertificate=yes"
                    f"&encrypt=no"
                )
            cls._engine = create_engine(
                 url,
                poolclass=QueuePool,
                pool_size=10,
                max_overflow=20,
                pool_timeout=30,
                pool_recycle=3600,
                pool_pre_ping=True,
                fast_executemany=True,
                echo=False,
                connect_args={
                    'application_name': 'Qagent_Local_App',
                    'timeout': 30
                }
            )
        return cls._engine

    async def execute(self, consulta: str) -> str:
        try:
            engine = self.get_engine()
            logger.info(f"Explicando consulta: {consulta}")

            conn = engine.raw_connection()
            try:
                cursor = conn.cursor()

                # Activar análisis del plan
                cursor.execute("SET SHOWPLAN_ALL ON;")
                conn.commit()

                # Ejecutar la consulta generada por IA (solo plan, no ejecuta)
                cursor.execute(consulta)
                columns = [col[0] for col in cursor.description]
                rows = cursor.fetchall()
                plan = [dict(zip(columns, row)) for row in rows]

                # Desactivar plan
                cursor.execute("SET SHOWPLAN_ALL OFF;")
                conn.commit()

            finally:
                conn.close()

            # Agregar instrucción al inicio
            plan.insert(0, {
                "Instrucción adicional": (
                    "⚠️ Esta respuesta contiene un plan de ejecución. "
                    "Razona muy bien si la consulta SQL es la adecuada y replantéala si lo estimas conveniente, "
                    "Si no son muchos registros **informa al usuario** cuantos registros contiene y que es posible que se demore un poco en obtener los resultado ya que replantearás la consulta."
                )
            })
            
            logger.info(f"Explicando consulta: {plan}")        
            return json.dumps(plan, indent=2, default=str)        
        
        except (SQLAlchemyError, pyodbc.Error) as err:
            error_message = str(err)
            logger.warning(f"1 Este es el error: {error_message}")
            
            response = get_random_response("error")
            await cl.Message(response).send()
            
            if "Invalid column name" in error_message:
                return f"> Este error corresponde al plan de ejecución de la consulta: {error_message}. -  Por favor revisa el nombre de las columnas en tu base de conocimiento y otorga la respuesta correcta"
            else:
                return f">  Este error corresponde al plan de ejecución de la consulta: {error_message}. -  Por favor replantea la consulta y otorga la respuesta correcta"
        except Exception as e:
            error_message = str(e)
            logger.warning(f"2 Este es el error: {error_message}")
            return f">  Este error corresponde al plan de ejecución de la consulta: {error_message}. -  Por favor replantea la consulta y otorga la respuesta correcta"

class GetExplainASQLS(BaseTool):
    """
    Herramienta para explicar  consultas Azure SQL SERVER
    """
    _engine = None

    @classmethod
    def get_engine(cls):
        if cls._engine is None:
            SERVER   = config.get("DB_ASQLS_SERVER")
            DATABASE = config.get("DB_ASQLS_DATABASE")
            USERNAME = config.get("DB_ASQLS_USERNAME")
            PASSWORD = config.get("DB_ASQLS_PASSWORD")
            DRIVER   = config.get("DB_ASQLS_DRIVER")

            url = (
                f"mssql+pyodbc://{USERNAME}:{PASSWORD}"
                f"@{SERVER}:1433/{DATABASE}"
                f"?driver={DRIVER.replace(' ', '+')}&TrustServerCertificate=yes&encrypt=yes"
            )
            
            
            cls._engine = create_engine(
                url,
                poolclass=QueuePool,
                pool_size=10,
                max_overflow=20,
                pool_timeout=30,
                pool_recycle=3600,
                pool_pre_ping=True,
                fast_executemany=True,
                echo=False,
                connect_args={
                    'application_name': 'Qagent_App',
                    'timeout': 30
                }
            )
        return cls._engine


    async def execute(self, consulta: str) -> str:
        try:
            engine = self.get_engine()
            logger.info(f"Explicando consulta: {consulta}")

            conn = engine.raw_connection()
            try:
                cursor = conn.cursor()

                # Activar análisis del plan
                cursor.execute("SET SHOWPLAN_ALL ON;")
                conn.commit()

                # Ejecutar la consulta generada por IA (solo plan, no ejecuta)
                cursor.execute(consulta)
                columns = [col[0] for col in cursor.description]
                rows = cursor.fetchall()
                plan = [dict(zip(columns, row)) for row in rows]

                # Desactivar plan
                cursor.execute("SET SHOWPLAN_ALL OFF;")
                conn.commit()

            finally:
                conn.close()

            # Agregar instrucción al inicio
            plan.insert(0, {
                "Instrucción adicional": (
                    "⚠️ Esta respuesta contiene un plan de ejecución. "
                    "Razona muy bien si la consulta SQL es la adecuada y replantéala si lo estimas conveniente, "
                    "Si no son muchos registros **informa al usuario** cuantos registros contiene y que es posible que se demore un poco en obtener los resultado ya que replantearás la consulta."
                )
            })
            
            logger.info(f"Respuesta: {plan}")        
            return json.dumps(plan, indent=2, default=str)        
        
        except (SQLAlchemyError, pyodbc.Error) as err:
            error_message = str(err)
            logger.warning(f"1 Este es el error: {error_message}")
            
            response = get_random_response("error")
            await cl.Message(response).send()
            
            if "Invalid column name" in error_message:
                return f"> Este error corresponde al plan de ejecución de la consulta: {error_message}. -  Por favor revisa el nombre de las columnas en tu base de conocimiento y otorga la respuesta correcta"
            else:
                return f">  Este error corresponde al plan de ejecución de la consulta: {error_message}. -  Por favor replantea la consulta y otorga la respuesta correcta"
        except Exception as e:
            error_message = str(e)
            logger.warning(f"2 Este es el error: {error_message}")
            return f">  Este error corresponde al plan de ejecución de la consulta: {error_message}. -  Por favor replantea la consulta y otorga la respuesta correcta"

class GetExplainGCPSQLS(BaseTool):
    """
    Herramienta para explicar  consultas SQL SERVER GCP
    """
    _engine = None

    @classmethod
    def get_engine(cls):
        if cls._engine is None:
            SERVER   = config.get("DB_GCP_SQLS_SERVER")
            DATABASE = config.get("DB_CP_SQLS_DATABASE")
            USERNAME = config.get("DB_DB_GCP_SQLS_USERNAME")
            PASSWORD = config.get("DB_GCP_SQLS_PASSWORD")
            DRIVER   = config.get("DB_GCP_SQLS_DRIVER")  # "ODBC Driver 18 for SQL Server"

            url = (
                f"mssql+pyodbc://{USERNAME}:{PASSWORD}"
                f"@{SERVER}:1433/{DATABASE}"
                f"?driver={DRIVER.replace(' ', '+')}"
                f"&TrustServerCertificate=yes"
                f"&encrypt=yes"
            )
            cls._engine = create_engine(
                url,
                poolclass=QueuePool,
                pool_size=10,
                max_overflow=20,
                pool_timeout=30,
                pool_recycle=3600,
                pool_pre_ping=True,
                fast_executemany=True,
                echo=False,
                connect_args={
                    'application_name': 'Qagent_GCP_App',
                    'timeout': 30
                }
            )
        return cls._engine

    async def execute(self, consulta: str) -> str:
        try:
            engine = self.get_engine()
            logger.info(f"Explicando consulta: {consulta}")

            conn = engine.raw_connection()
            try:
                cursor = conn.cursor()

                # Activar análisis del plan
                cursor.execute("SET SHOWPLAN_ALL ON;")
                conn.commit()

                # Ejecutar la consulta generada por IA (solo plan, no ejecuta)
                cursor.execute(consulta)
                columns = [col[0] for col in cursor.description]
                rows = cursor.fetchall()
                plan = [dict(zip(columns, row)) for row in rows]

                # Desactivar plan
                cursor.execute("SET SHOWPLAN_ALL OFF;")
                conn.commit()

            finally:
                conn.close()

            # Agregar instrucción al inicio
            plan.insert(0, {
                "Instrucción adicional": (
                    "⚠️ Esta respuesta contiene un plan de ejecución. "
                    "Razona muy bien si la consulta SQL es la adecuada y replantéala si lo estimas conveniente, "
                    "Si no son muchos registros **informa al usuario** cuantos registros contiene y que es posible que se demore un poco en obtener los resultado ya que replantearás la consulta."
                )
            })
            
            logger.info(f"Explicando consulta: {plan}")        
            return json.dumps(plan, indent=2, default=str)        
        
        except (SQLAlchemyError, pyodbc.Error) as err:
            error_message = str(err)
            logger.warning(f"1 Este es el error: {error_message}")
            
            response = get_random_response("error")
            await cl.Message(response).send()
            
            if "Invalid column name" in error_message:
                return f"> Este error corresponde al plan de ejecución de la consulta: {error_message}. -  Por favor revisa el nombre de las columnas en tu base de conocimiento y otorga la respuesta correcta"
            else:
                return f">  Este error corresponde al plan de ejecución de la consulta: {error_message}. -  Por favor replantea la consulta y otorga la respuesta correcta"
        except Exception as e:
            error_message = str(e)
            logger.warning(f"2 Este es el error: {error_message}")
            return f">  Este error corresponde al plan de ejecución de la consulta: {error_message}. -  Por favor replantea la consulta y otorga la respuesta correcta"

#única activa
@cl.step(type="tool")
async def getdataMSQL(consulta: str) -> str:
    """
    Herramienta de Assistant
    Genera la respuesta por medio de una consulta SQL a la base de datos.

    Args:
        consulta (str): consulta del usuario para crear la consulta sql 
            
    Returns: 
        Retorna una instrucción de tipo dict al asistente
    """
    #tool = GetDataLocalMySQLTool()
    tool = GetDataMySQLAsync()
    return await tool.execute(sql=consulta)

async def getdataASQLS(consulta: str) -> str:
    """
    Herramienta de Assistant
    Genera la respuesta por medio de una consulta SQL a la base de datos Azure SQL Server.

    Args:
        consulta (str): consulta del usuario para crear la consulta sql 
            
    Returns: 
        Retorna una instrucción de tipo dict al asistente
    """
    tool = GetDataAzureSQLServer()
    return await tool.execute(consulta=consulta)

async def getdataASQLS_AD(consulta: str) -> str:
    """
    Herramienta de Assistant
    Genera la respuesta por medio de una consulta SQL a la base de datos Azure SQL Server.

    Args:
        consulta (str): consulta del usuario para crear la consulta sql 
            
    Returns: 
        Retorna una instrucción de tipo dict al asistente
    """
    tool = GetDataAzureSQLServerAD()
    return await tool.execute(consulta=consulta)

async def getdataGSQLS(consulta: str) -> str:
    """
    Herramienta de Assistant
    Genera la respuesta por medio de una consulta SQL a la base de datos Azure SQL Server.

    Args:
        consulta (str): consulta del usuario para crear la consulta sql 
            
    Returns: 
        Retorna una instrucción de tipo dict al asistente
    
    Comentario:    
    Puerto 1433 debe estar abierto y permitido para IP o red.    
    """
    tool = GetDataGCPSQLServer()
    return await tool.execute(consulta=consulta)

async def getdataSQLSLocal(consulta: str) -> str:
    """
    Herramienta de Assistant
    Genera la respuesta por medio de una consulta SQL a la base de datos Azure SQL Server.

    Args:
        consulta (str): consulta del usuario para crear la consulta sql 
            
    Returns: 
        Retorna una instrucción de tipo dict al asistente
    
    Comentario:    
    Puerto 1433 debe estar abierto y permitido para IP o red.    
    """
    tool = GetDataSQLSLocal()
    return await tool.execute(consulta=consulta)

async def explainSQL(consulta: str) -> str:
    """
    Herramienta de Assistant
   Explica la consulta sql para mejorar la respuesta del modelo 

    Args:
        consulta (str): consulta del usuario para crear la consulta sql 
            
    Returns: 
        Retorna la explicación de la consulta
    
    Example:
    
        {
            "StmtText": "SELECT * FROM tabla;",
            "StmtId": 1,
            "NodeId": 0,
            "Parent": NULL,
            "PhysicalOp": "Clustered Index Scan",
            "LogicalOp": "Clustered Index Scan",
            "EstimateRows": 34118.0,
            ...
        }
    
    """
    tool = GetExplainSQL()
    return await tool.execute(consulta=consulta)

async def explainGCPSQL(consulta: str) -> str:
    """
    Herramienta de Assistant
   Explica la consulta sql para mejorar la respuesta del modelo 

    Args:
        consulta (str): consulta del usuario para crear la consulta sql 
            
    Returns: 
        Retorna la explicación de la consulta
    
    Example:
    
        {
            "StmtText": "SELECT * FROM tabla;",
            "StmtId": 1,
            "NodeId": 0,
            "Parent": NULL,
            "PhysicalOp": "Clustered Index Scan",
            "LogicalOp": "Clustered Index Scan",
            "EstimateRows": 34118.0,
            ...
        }
    
    """
    tool = GetExplainGCPSQLS()
    return await tool.execute(consulta=consulta)

async def explainASQL(consulta: str) -> str:
    """
    Herramienta de Assistant
   Explica la consulta sql para mejorar la respuesta del modelo 

    Args:
        consulta (str): consulta del usuario para crear la consulta sql 
            
    Returns: 
        Retorna la explicación de la consulta
    
    Example:
    
        {
            "StmtText": "SELECT * FROM tabla;",
            "StmtId": 1,
            "NodeId": 0,
            "Parent": NULL,
            "PhysicalOp": "Clustered Index Scan",
            "LogicalOp": "Clustered Index Scan",
            "EstimateRows": 34118.0,
            ...
        }
    
    """
    tool = GetExplainASQLS()
    return await tool.execute(consulta=consulta)

async def createDataFrame(message, dataframe: str) -> str:
    """
    Herramienta de Assistant
    Genera la respuesta por medio de una consulta SQL a la base de datos Azure SQL Server.

    Args:
        consulta (str): consulta del usuario para crear la consulta sql 
            
    Returns: 
        Retorna una instrucción de tipo dict al asistente
    """
    tool = CreateDataFrameTool()
    return await tool.execute(message=message, dataframe=dataframe)

async def generar_excel_desde_mysql(resultados):
    
    import openpyxl
    from openpyxl.utils import get_column_letter
    import os
    from datetime import datetime

    try:
        # Crear libro de Excel
        libro = openpyxl.Workbook()
        hoja = libro.active
        
        logger.info(f"Generar_excel_desde_mysql:   {resultados}")

        # Obtener nombres de columnas (claves del primer diccionario)
        columnas = list(resultados[0].keys())
        # Escribir encabezados
        for col_num, columna in enumerate(columnas, 1):
            hoja.cell(row=1, column=col_num, value=columna)
        
        # Escribir datos
        for row_num, fila in enumerate(resultados, 2):
            for col_num, columna in enumerate(columnas, 1):
                valor = fila[columna]
                
                # Manejar tipos de datos especiales (como datetime)
                if hasattr(valor, 'isoformat'):
                    valor = valor.isoformat()
                
                hoja.cell(row=row_num, column=col_num, value=valor)
        
        # Ajustar ancho de columnas automáticamente
        for col_num, columna in enumerate(columnas, 1):
            # Longitud del encabezado
            longitudes = [len(str(columna))]

            # Longitud de los valores no nulos
            for fila in resultados:
                valor = fila.get(columna)
                if valor is not None:
                    longitudes.append(len(str(valor)))

            hoja.column_dimensions[get_column_letter(col_num)].width = min(max(longitudes) + 2, 50)
        
        # Generar nombre de archivo si no se proporcionó
       
        fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"public/storage/excel/reporte_{fecha_actual}.xlsx"
        libro.save(nombre_archivo)
        URL_ARCHIVO = config.get("URL_ARCHIVO")
        url_archivo = f"{URL_ARCHIVO}{nombre_archivo}"
        logger.info(f"url_archivo:   {url_archivo}")
        
        
        return url_archivo
    except Exception as err:
         logger.error(f"Error al crear un excel: {err}")