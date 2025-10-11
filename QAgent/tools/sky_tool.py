# -*- coding: utf-8 -*-
"""
Herramienta de auditoría ISARP (solo payload).
La IA debe llamar: await audit_excel_tool(payload)
"""

from __future__ import annotations
import json
import logging
from datetime import datetime
from typing import List, Dict, Optional, Union

import chainlit as cl
from openpyxl import load_workbook

# Importa tu BaseTool y utilidades reales de tu proyecto
from QAgent.tools.base_tool import BaseTool
from QAgent.utils.logging_utils import notify_error, get_random_response

logger = logging.getLogger(__name__)

# ---------------- Encabezados esperados ----------------
HEADER_ISARP = "ISARP"
HEADER_DOC_REFS = "Documentation References (ACTUALIZAR USANDO NUEVO MANUAL)"
HEADER_RESULT = "Resultado Auditoria Documental (CONFORME / NO CONFORME)"
HEADER_JUST = "Description of Nonconformity or Description of Reason for N/A"
VALID_RESULTS = {"CONFORME", "NO CONFORME", "N/A"}

# ---------------- Helpers internos ----------------
def _build_header_map(ws) -> Dict[str, int]:
    header_map = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        title = (cell.value or "").strip()
        if title:
            header_map[title] = col_idx
    missing = {HEADER_ISARP, HEADER_DOC_REFS, HEADER_RESULT, HEADER_JUST} - set(header_map.keys())
    if missing:
        raise ValueError(f"Faltan columnas en el encabezado: {', '.join(missing)}")
    return header_map

def _index_rows_by_isarp(ws, isarp_col: int) -> Dict[str, int]:
    idx = {}
    for row in range(2, ws.max_row + 1):
        key = ws.cell(row=row, column=isarp_col).value
        if key is not None:
            key_str = str(key).strip()
            if key_str and key_str not in idx:
                idx[key_str] = row
    return idx

def _apply_item(ws, header_map, row: int,
               documentation_references: str,
               resultado_auditoria: str,
               justificacion: str) -> None:
    res_norm = (resultado_auditoria or "").strip().upper()
    if res_norm not in VALID_RESULTS:
        raise ValueError(
            f"resultado_auditoria inválido: {res_norm}. "
            f"Permitidos: {', '.join(sorted(VALID_RESULTS))}"
        )
    ws.cell(row=row, column=header_map[HEADER_DOC_REFS]).value = documentation_references
    ws.cell(row=row, column=header_map[HEADER_RESULT]).value = res_norm
    ws.cell(row=row, column=header_map[HEADER_JUST]).value = justificacion

def _ensure_payload_list(payload: Union[str, List[Dict[str, str]]]) -> Union[List[Dict[str, str]], str]:
    """Valida el payload. Devuelve lista válida o un string con el mensaje de error."""
    if isinstance(payload, str):
        try:
            data = json.loads(payload)
        except json.JSONDecodeError as e:
            return f"⚠️ Payload no es JSON válido: {e}"
    else:
        data = payload

    if not isinstance(data, list):
        return "⚠️ Payload debe ser una lista de objetos."

    required = ("isarp", "documentation_references", "resultado_auditoria", "justificacion")
    for i, it in enumerate(data, 1):
        if not isinstance(it, dict):
            return f"⚠️ Payload[{i}] no es un objeto/dict."
        for k in required:
            if k not in it:
                if k == "justificacion":
                    return "⚠️ Te falta la justificación piensa mejor la respuesta e intenta nuevamente."
                return f"⚠️ Payload[{i}] carece del campo requerido '{k}'."
            # No vacío
            if isinstance(it[k], str) and not it[k].strip():
                if k == "justificacion":
                    return "⚠️ Te falta la justificación piensa mejor la respuesta e intenta nuevamente."
                return f"⚠️ Payload[{i}].{k} no puede estar vacío."
    return data


# ---------------- Herramienta ----------------
class AuditExcelTool(BaseTool):
    """
    Herramienta para escribir resultados de auditoría en Excel.
    Se usa SOLO en modo batch: recibe un único parámetro `payload` (list[dict] o JSON string).
    Otros parámetros quedan fijos aquí y luego los moverás a .env.
    """

    async def execute(
        self,
        payload: Union[str, List[Dict[str, str]]]
    ) -> str:
        # ---------------- Parámetros fijos (migrarán a .env) ----------------
        logger.info("============ creando los parámetros ============")
        
        xlsx_path = "public/storage/sky.xlsx"
        sheet_name = "Conformance Report"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        save_as = f"public/storage/{timestamp}_sky_report.xlsx"
        url_archivo = f"http://localhost:8000/public/storage/{timestamp}_sky_report.xlsx"        
        strict = False
        
        logger.info("============ PARAMETROS ============")
        logger.info(
                f"\n=============================\n"
                f"  xlsx_path : {xlsx_path}\n"
                f"  save_as : {save_as}\n"
                f"  Link     : {url_archivo}\n"
                f"\n=============================\n"
            )
        
        
        # -------------------------------------------------------------------

        try:
            # Validar payload
            items = _ensure_payload_list(payload)
            if isinstance(items, str): 
                return items
            if not items:
                return "⚠️ Payload vacío: no hay items para procesar."

            wb = load_workbook(xlsx_path)
            ws = wb[sheet_name] if sheet_name else wb.active
            header_map = _build_header_map(ws)
            idx = _index_rows_by_isarp(ws, header_map[HEADER_ISARP])

            ok, skipped = 0, []
            for it in items:
                isarp = str(it.get("isarp", "")).strip()
                if not isarp:
                    if strict:
                        raise ValueError("Falta 'isarp' en uno de los items.")
                    skipped.append("item-sin-isarp")
                    continue
                if isarp not in idx:
                    if strict:
                        raise KeyError(f"No se encontró una fila con ISARP='{isarp}'.")
                    skipped.append(isarp)
                    continue

                row = idx[isarp]
                _apply_item(
                    ws, header_map, row,
                    documentation_references=it.get("documentation_references", ""),
                    resultado_auditoria=it.get("resultado_auditoria", ""),
                    justificacion=it.get("justificacion", "")
                )
                ok += 1

            wb.save(save_as)
            detalle_omitidos = "-" if not skipped else ", ".join(skipped)
            
           
            
            logger.info(
                f"\n=============================\n"
                f"✅ Escritura batch completada.\n"
                f"   OK       : {ok}\n"
                f"   OMITIDOS : {len(skipped)} ({detalle_omitidos})\n"
                f"   Archivo  : {save_as}"
                f"   Link     : {url_archivo}"
                f"\n=============================\n"
            )
            
            return (
                f"✅ Escritura batch completada.\n"
                f"   OK       : {ok}\n"
                f"   OMITIDOS : {len(skipped)} ({detalle_omitidos})\n"
                f"   Archivo  : {save_as}"
                f"   Despiega un link en md del reporte  [Descargar Excel]({url_archivo})"
                f"   Resume tu análisis detalladamente"
            )

        except Exception as err:
            # Log + feedback al usuario
            await notify_error(str(err), "audit_excel", "AuditExcelTool.execute")
            response = get_random_response("error")
            await cl.Message(response).send()
            return f"❌ Error en AuditExcelTool: {err}"

# ---------------- Wrapper tipo @cl.step ----------------
@cl.step(type="tool")
async def audit_excel_tool(
    payload: Union[str, List[Dict[str, str]]]
) -> str:
    """
    Escribe resultados de auditoría ISARP en Excel.
    Recibe UN SOLO parámetro: `payload`
    - Puede ser lista de dicts o un string JSON con el siguiente formato:

    [
      {
        "isarp": "17-ORG 2.1.1",
        "documentation_references": "Manual ISM 17 §2.1.1; Proced. OP-12",
        "resultado_auditoria": "CONFORME",
        "justificacion": "Evidencia de registros actualizados al 2025-08-20."
      },
      ...
    ]
    """
    tool = AuditExcelTool()
    return await tool.execute(payload=payload)



"""
{
  "name": "audit_excel_tool",
  "description": " Escribe resultados de auditoría ISARP en Excel.Recibe UN SOLO parámetro: `payload`.",
  "parameters": {
    "type": "object",
    "required": [
      "payload"
    ],
    "properties": {
      "payload": {
        "type": "array",
        "items": {
          "type": "string"
        },
        "description": "Lista de objetos con resultados de auditoría ISARP. Ejemplo: [{isarp:17-ORG 2.1.1,documentation_references:Manual ISM 17 §2.1.1; Proced. OP-12,resultado_auditoria:CONFORME,justificacion:Evidencia de registros actualizados al 2025-08-20.},{isarp:17-ORG 2.1.4,documentation_references:Manual ISM 17 §2.1.4,resultado_auditoria:NO CONFORME,justificacion:Falta evidencia de capacitación anual en seguridad.}]"
      }
    }
  }
}


"""