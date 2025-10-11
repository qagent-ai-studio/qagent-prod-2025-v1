# test_audit_excel_fixed.py
# -*- coding: utf-8 -*-
from __future__ import annotations
from typing import List, Dict, Optional
from openpyxl import load_workbook
from datetime import datetime

# ---------------- Encabezados esperados ----------------
HEADER_ISARP = "ISARP"
HEADER_DOC_REFS = "Documentation References (ACTUALIZAR USANDO NUEVO MANUAL)"
HEADER_RESULT = "Resultado Auditoria Documental (CONFORME / NO CONFORME)"
HEADER_JUST = "Description of Nonconformity or Description of Reason for N/A"
VALID_RESULTS = {"CONFORME", "NO CONFORME", "N/A"}

def _build_header_map(ws) -> Dict[str, int]:
    header_map = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        title = (cell.value or "").strip()
        if title:
            header_map[title] = col_idx
    missing = {HEADER_ISARP, HEADER_DOC_REFS, HEADER_RESULT, HEADER_JUST} - set(header_map.keys())
    if missing:
        raise ValueError(f"Faltan columnas en el encabezado: {', '.join(missing)}")
    return header_map

def _index_rows_by_isarp(ws, isarp_col: int) -> Dict[str, int]:
    idx = {}
    for row in range(2, ws.max_row + 1):
        key = ws.cell(row=row, column=isarp_col).value
        if key is not None:
            key_str = str(key).strip()
            if key_str and key_str not in idx:
                idx[key_str] = row
    return idx

def _apply_item(ws, header_map, row: int,
               documentation_references: str,
               resultado_auditoria: str,
               justificacion: str) -> None:
    res_norm = (resultado_auditoria or "").strip().upper()
    if res_norm not in VALID_RESULTS:
        raise ValueError(f"resultado_auditoria inválido: {res_norm}")
    ws.cell(row=row, column=header_map[HEADER_DOC_REFS]).value = documentation_references
    ws.cell(row=row, column=header_map[HEADER_RESULT]).value = res_norm
    ws.cell(row=row, column=header_map[HEADER_JUST]).value = justificacion

def apply_audit_results_batch(
    xlsx_path: str,
    items: List[Dict[str, str]],
    sheet_name: Optional[str] = None,
    save_as: Optional[str] = None,
    strict: bool = False
) -> Dict[str, str]:
    if not items:
        return {"ok": "0", "skipped": "no-items"}

    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    header_map = _build_header_map(ws)
    isarp_col = header_map[HEADER_ISARP]
    idx = _index_rows_by_isarp(ws, isarp_col)

    ok_count = 0
    skipped = []

    for it in items:
        isarp = str(it.get("isarp", "")).strip()
        if not isarp:
            if strict:
                raise ValueError("Falta 'isarp' en uno de los items.")
            skipped.append("item-sin-isarp")
            continue

        if isarp not in idx:
            if strict:
                raise KeyError(f"No se encontró una fila con ISARP='{isarp}'.")
            skipped.append(isarp)
            continue

        row = idx[isarp]
        _apply_item(
            ws, header_map, row,
            documentation_references=it.get("documentation_references", ""),
            resultado_auditoria=it.get("resultado_auditoria", ""),
            justificacion=it.get("justificacion", "")
        )
        ok_count += 1

    wb.save(save_as or xlsx_path)
    return {"ok": str(ok_count), "skipped": ", ".join(skipped)}

# ---------------- MAIN ----------------
def main():
    # Parámetros fijos
    xlsx_path = "public/storage/sky.xlsx"
    sheet_name = "Conformance Report"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    save_as = f"public/storage/{timestamp}_sky_report.xlsx"
    strict = False

    # Payload de prueba
    payload = [
      {
        "isarp": "17-ORG 2.1.1",
        "documentation_references": "Manual ISM 17 §2.1.1; Proced. OP-12",
        "resultado_auditoria": "CONFORME",
        "justificacion": "Evidencia de registros actualizados al 2025-08-20."
      },
      {
        "isarp": "17-ORG 2.1.4",
        "documentation_references": "Manual ISM 17 §2.1.4",
        "resultado_auditoria": "NO CONFORME",
        "justificacion": "Falta evidencia de capacitación anual en seguridad."
      }
    ]

    try:
        summary = apply_audit_results_batch(
            xlsx_path=xlsx_path,
            items=payload,
            sheet_name=sheet_name,
            save_as=save_as,
            strict=strict
        )
        print("✅ Escritura completada.")
        print(f"   OK      : {summary['ok']}")
        print(f"   OMITIDOS: {summary['skipped'] or '-'}")
        print(f"   Archivo : {save_as}")
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    main()
